"""
generate_cargo_template.py

화물 위치를 엑셀로 한번에 등록할 수 있는 입력 양식(cargo_template.xlsx)을 생성합니다.
등록된 위치 목록(location_marks_verified.json / location_marks.json)이 있으면 자동으로
불러와서 "현재위치" 컬럼에 드롭다운으로 넣어줍니다 (오타 방지).

실행:
    python generate_cargo_template.py
    -> cargo_template.xlsx 생성됨
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# 실행 위치(cwd)와 무관하게 항상 이 스크립트가 있는 폴더의 파일을 쓰도록 고정합니다.
_APP_DIR = Path(__file__).resolve().parent
VERIFIED_LOCATIONS_FILE = str(_APP_DIR / "location_marks_verified.json")
SIMPLE_LOCATIONS_FILE = str(_APP_DIR / "location_marks.json")
OUTPUT_FILE = str(_APP_DIR / "cargo_template.xlsx")

HEADER_ROW = 4
EXAMPLE_ROW = 5
DATA_START_ROW = 6
DATA_END_ROW = 205  # 드롭다운을 적용할 입력 가능 행 범위 (약 200행)

FONT_NAME = "Arial"


def load_known_locations():
    if Path(VERIFIED_LOCATIONS_FILE).exists():
        data = json.loads(Path(VERIFIED_LOCATIONS_FILE).read_text(encoding="utf-8"))
        return list(data.keys())
    if Path(SIMPLE_LOCATIONS_FILE).exists():
        data = json.loads(Path(SIMPLE_LOCATIONS_FILE).read_text(encoding="utf-8"))
        return list(data.keys())
    return []


def build_template() -> None:
    known_locations = load_known_locations()  # 드롭다운에 넣을 위치 이름 목록 미리 확보

    wb = Workbook()          # 새 엑셀 파일(워크북) 생성
    ws = wb.active           # 기본으로 생성되는 첫 번째 시트를 가져옴
    ws.title = "화물등록"     # 시트 이름 지정

    # ---- 제목 ----
    ws.merge_cells("A1:E1")  # A1~E1을 하나로 합쳐서 제목이 여러 칸에 걸쳐 보이게 함
    ws["A1"] = "화물 위치 일괄 등록 양식"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True)

    # ---- 안내문 ----
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "아래 표에 화물명과 현재위치를 입력하세요. '현재위치'는 드롭다운에서 선택할 수 있습니다"
        " (목록은 '등록된위치목록' 시트 참고). 노란색 예시 행(5행)은 삭제하거나 덮어쓰고 사용하세요."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")  # 긴 문장이 잘리지 않고 줄바꿈되게
    ws.row_dimensions[2].height = 30  # 안내문이 2줄 정도 나올 수 있게 행 높이를 키움

    # ---- 헤더 (4행) ----
    headers = ["화물명", "현재위치", "컨테이너/ArUco ID", "화물종류", "비고"]
    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")  # 파란색 배경
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=title)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")  # 흰색 굵은 글씨
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ---- 예시 행 (5행, 노란색으로 구분해서 실제 데이터가 아님을 알려줌) ----
    example_values = ["화물A", known_locations[0] if known_locations else "창고", "ARUCO_12", "컨테이너", "예시 행입니다 - 삭제 후 사용하세요"]
    example_fill = PatternFill(start_color="FFF6B3", end_color="FFF6B3", fill_type="solid")
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=EXAMPLE_ROW, column=col_idx, value=value)
        cell.font = Font(name=FONT_NAME, size=11, italic=True, color="7A6A00")
        cell.fill = example_fill

    # ---- 컬럼 너비 (읽기 좋게 각 컬럼 폭 지정) ----
    widths = [16, 16, 18, 14, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width  # 64+1='A', 64+2='B' ... 아스키코드로 열 문자 생성

    # ---- 등록된 위치 목록 시트 (드롭다운이 참조할 원본 데이터 + 사람이 봐도 참고용) ----
    ws_loc = wb.create_sheet("등록된위치목록")  # 두 번째 시트 새로 생성
    ws_loc["A1"] = "위치명"
    ws_loc["A1"].font = Font(name=FONT_NAME, size=11, bold=True)
    ws_loc.column_dimensions["A"].width = 20

    if known_locations:
        for i, name in enumerate(known_locations, start=2):  # 2행부터 (1행은 헤더)
            ws_loc.cell(row=i, column=1, value=name)
    else:
        ws_loc["A2"] = "(등록된 위치 없음 - location_marks.json 생성 후 다시 만들어주세요)"
        ws_loc["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="EA5455")

    # ---- 데이터 검증(드롭다운) : "현재위치" 컬럼(B열)에 적용 ----
    if known_locations:
        last_row = len(known_locations) + 1  # 위치 목록 시트에서 데이터가 있는 마지막 행 번호
        dv = DataValidation(
            type="list",  # 목록에서 선택하는 방식의 검증
            formula1=f"등록된위치목록!$A$2:$A${last_row}",  # 드롭다운 항목을 다른 시트 범위에서 가져옴
            allow_blank=True,
            showDropDown=False,  # openpyxl 특성상 False로 둬야 실제로 드롭다운 화살표가 보임 (반직관적이지만 라이브러리 버그성 동작)
            showErrorMessage=True,  # 목록에 없는 값 입력 시 경고창이 뜨게 함
        )
        dv.error = "등록된 위치 목록에 없는 이름입니다. '등록된위치목록' 시트를 확인해주세요."
        dv.errorTitle = "위치명 확인 필요"
        ws.add_data_validation(dv)          # 검증 규칙을 시트에 등록
        dv.add(f"B{DATA_START_ROW}:B{DATA_END_ROW}")  # 6행~205행의 B열 전체에 이 규칙 적용

    wb.save(OUTPUT_FILE)  # 지금까지 만든 내용을 실제 .xlsx 파일로 저장
    print(f"생성 완료: {OUTPUT_FILE}")
    print(f"등록된 위치 {len(known_locations)}개를 드롭다운에 반영했습니다: {known_locations}")


if __name__ == "__main__":
    build_template()

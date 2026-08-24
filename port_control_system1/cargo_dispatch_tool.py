"""
cargo_dispatch_tool.py

화물이 지금 어느 등록된 위치에 있는지 기록해두는 도구입니다.
  1) 화물 위치를 수동/엑셀 일괄 등록
  2) 위치가 바뀐 화물이 담긴 엑셀을 불러오면, 예전 위치와 다른 화물만 골라
     [대기장소 -> 픽업 -> 새 위치] 경로를 자동 생성해서 큐로 순서대로 실행 (데모)
  3) 완료되면 화물의 위치를 목적지로 갱신

자연어 명령("화물A를 항구로 옮겨")으로 배차하는 기능은 command_center.py(우측 하단 "🗣️ 명령"
버튼 팝업)로 옮겼습니다 - 이 파일의 build_route/RouteStep/load_cargo_registry 등을 그대로
가져다 씁니다.

location_marks_verified.json (또는 location_marks.json)에 이미 등록해두신 위치들을
그대로 재사용합니다. 실차 연동 시에는 파일 하단의 CargoDispatcherROS2 참고 코드로
Nav2 목표 전송 + 적재/하역 확인 신호 대기를 연결하시면 됩니다.
"""

import json
import os
import queue
import threading
import time
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog
from typing import Dict, List, Optional, Tuple

import customtkinter as ctk
from openpyxl import load_workbook

from waypoint_rules import expand_leg, load_waypoint_rules
from generate_cargo_template import build_template as build_cargo_excel_template
from inventory_client import InventoryAdminClient, InventoryClient
from autonomous_inventory import (
    CANONICAL_LOCATIONS,
    SHIP_MARKERS,
    WAREHOUSE_MARKERS,
)


# 실행할 때의 현재 폴더(cwd)가 아니라, 이 파일이 실제로 있는 폴더를 기준으로 삼습니다.
# 예전에는 "cargo_locations.json"처럼 상대경로로 열고 저장했는데, 이러면 파이썬을
# 어느 폴더에서 실행하느냐에 따라 실제로 읽고 쓰는 파일이 달라지는 문제가 있었습니다
# (예: Downloads에서 실행할 때와 프로젝트 폴더에서 실행할 때 서로 다른 파일을 봄).
# 이제는 항상 이 스크립트가 있는 폴더의 파일을 쓰도록 고정합니다.
_APP_DIR = Path(__file__).resolve().parent

VERIFIED_LOCATIONS_FILE = str(_APP_DIR / "location_marks_verified.json")
SIMPLE_LOCATIONS_FILE = str(_APP_DIR / "location_marks.json")
CARGO_FILE = str(_APP_DIR / "cargo_locations.json")
STANDBY_LOCATION = "대기장소"  # 등록된 개별 대기 위치가 없을 때 쓰는 예비 기본값

# ---- 차량 대수/위치 추적 (자연어 명령 팝업과 엑셀 재배치가 공용으로 씁니다) ----
NUM_VEHICLES = 2  # 보유 차량 대수 - 늘리거나 줄이려면 이 값만 바꾸면 됩니다.

# 차량마다 자기만의 "집(대기 위치)"이 따로 있는 경우 여기에 순서대로 적어주세요.
# 예: 차량1은 "대기장소 1"에서, 차량2는 "대기장소 2"에서 대기 - 실제 등록된 위치명과
# 글자까지 정확히 같아야 합니다. 이 목록이 NUM_VEHICLES보다 짧으면, 모자란 차량은
# STANDBY_LOCATION(공용 기본값)을 씁니다.
VEHICLE_HOME_LOCATIONS = ["대기장소 1", "대기장소 2"]

VEHICLE_STATE_FILE = str(_APP_DIR / "vehicle_positions.json")

MANUAL_INVENTORY_LOCATIONS = tuple(CANONICAL_LOCATIONS)
MANUAL_CONTAINER_IDS = tuple(str(value) for value in range(9))
MANUAL_FLOORS = ('1', '2', '3')
TRAILER_BASE_MARKERS = {'AMR1': '10', 'AMR2': '9'}


def validate_manual_inventory_position(
    container_id: str,
    location: str,
    floor: int,
    cargos,
) -> str:
    """Validate a manual DB edit and return its supporting ArUco/ID."""
    container_id = str(container_id).strip()
    location = str(location).strip()
    try:
        floor = int(floor)
    except (TypeError, ValueError) as exc:
        raise ValueError('층수는 1, 2, 3 중 하나여야 합니다.') from exc
    if container_id not in MANUAL_CONTAINER_IDS:
        raise ValueError('컨테이너 ID는 0부터 8까지 선택해주세요.')
    if location not in MANUAL_INVENTORY_LOCATIONS:
        raise ValueError(f'등록되지 않은 표준 위치입니다: {location}')
    if floor not in {1, 2, 3}:
        raise ValueError('층수는 1, 2, 3 중 하나여야 합니다.')
    if location in {'AMR1', 'AMR2', '출항완료'} and floor != 1:
        raise ValueError(f'{location} 위치는 1층만 사용할 수 있습니다.')

    records = tuple(cargos or ())
    for cargo in records:
        other_id = str(getattr(cargo, 'container_id', '') or '')
        if other_id == container_id:
            continue
        if (
            str(getattr(cargo, 'location', '') or '') == location
            and int(getattr(cargo, 'floor', 1) or 1) == floor
        ):
            raise ValueError(
                f'{location} {floor}층에는 이미 컨테이너 '
                f'{other_id or getattr(cargo, "name", "-")}가 있습니다.'
            )

    if floor == 1:
        return str(
            WAREHOUSE_MARKERS.get(location)
            or SHIP_MARKERS.get(location)
            or TRAILER_BASE_MARKERS.get(location)
            or ''
        )

    support = next((
        cargo for cargo in records
        if str(getattr(cargo, 'container_id', '') or '') != container_id
        and str(getattr(cargo, 'location', '') or '') == location
        and int(getattr(cargo, 'floor', 1) or 1) == floor - 1
    ), None)
    if support is None:
        raise ValueError(
            f'{location} {floor}층을 등록하려면 {floor - 1}층 '
            '컨테이너가 먼저 DB에 있어야 합니다.'
        )
    return str(getattr(support, 'container_id', '') or '')


def _inventory_refresh_interval_ms() -> int:
    """Return the dashboard DB polling interval with a safe lower bound."""
    try:
        seconds = float(os.environ.get("PORT_INVENTORY_UI_REFRESH_SEC", "2"))
    except (TypeError, ValueError):
        seconds = 2.0
    return max(500, int(max(0.5, seconds) * 1000))


def vehicle_home_location(vehicle_idx: int) -> str:
    """vehicle_idx번 차량의 기본 대기 위치 이름을 돌려줍니다."""
    if 0 <= vehicle_idx < len(VEHICLE_HOME_LOCATIONS):
        return VEHICLE_HOME_LOCATIONS[vehicle_idx]
    return STANDBY_LOCATION


def load_vehicle_state(num_vehicles: int = NUM_VEHICLES) -> Tuple[List[str], int]:
    """차량별 현재 위치와, 다음에 배차할 차량 순번을 불러옵니다.
    처음 실행하거나 파일이 없으면 각 차량이 자기 대기 위치(VEHICLE_HOME_LOCATIONS)에
    있는 것으로 시작합니다."""
    positions: List[str] = []
    next_index = 0

    if Path(VEHICLE_STATE_FILE).exists():
        try:
            data = json.loads(Path(VEHICLE_STATE_FILE).read_text(encoding="utf-8"))
            positions = list(data.get("positions", []))
            next_index = int(data.get("next_index", 0))
        except Exception:
            positions, next_index = [], 0

    if len(positions) < num_vehicles:
        for i in range(len(positions), num_vehicles):
            positions.append(vehicle_home_location(i))
    positions = positions[:num_vehicles]
    next_index = next_index % num_vehicles

    return positions, next_index


def save_vehicle_state(positions: List[str], next_index: int) -> None:
    Path(VEHICLE_STATE_FILE).write_text(
        json.dumps({"positions": positions, "next_index": next_index}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ---- 차량/로봇팔 상태 추적 (배터리, 이상 유무, 최근 작업 이력) ----
# 참고: 실제 센서 연동 전이라 배터리/이상 유무는 자동으로 갱신되지 않고, 기본값에서
# 시작해서 화면의 버튼으로 직접 표시를 바꾸는 방식입니다 (수동 관리).
# "마지막 작업"만 화물 이동을 실행할 때마다 자동으로 기록됩니다.
VEHICLE_STATUS_FILE = str(_APP_DIR / "vehicle_status.json")
ROBOT_ARM_STATUS_FILE = str(_APP_DIR / "robot_arm_status.json")


def _default_status() -> Dict:
    return {"battery_pct": 100, "fault": None, "last_job": None, "last_job_time": None}


def load_vehicle_status(num_vehicles: int = NUM_VEHICLES) -> Dict[str, Dict]:
    """차량별 상태를 불러옵니다. 파일이 없거나 새로 늘어난 차량은 기본값으로 채웁니다."""
    data = {}
    if Path(VEHICLE_STATUS_FILE).exists():
        try:
            data = json.loads(Path(VEHICLE_STATUS_FILE).read_text(encoding="utf-8"))
        except Exception:
            data = {}
    result = {}
    for i in range(num_vehicles):
        key = f"차량 {i + 1}"
        result[key] = {**_default_status(), **data.get(key, {})}
    return result


def save_vehicle_status(status: Dict[str, Dict]) -> None:
    Path(VEHICLE_STATUS_FILE).write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")


def record_vehicle_job(vehicle_idx: int, description: str) -> None:
    """화물 이동이 끝날 때마다 호출해서 그 차량의 "마지막 작업" 기록을 남깁니다."""
    status = load_vehicle_status()
    key = f"차량 {vehicle_idx + 1}"
    if key not in status:
        status[key] = _default_status()
    status[key]["last_job"] = description
    status[key]["last_job_time"] = time.strftime("%H:%M:%S")
    save_vehicle_status(status)


def load_robot_arm_status(names: List[str]) -> Dict[str, Dict]:
    """등록된 로봇팔 이름들의 상태를 불러옵니다. 새로 등록된 로봇팔은 기본값으로 채웁니다."""
    data = {}
    if Path(ROBOT_ARM_STATUS_FILE).exists():
        try:
            data = json.loads(Path(ROBOT_ARM_STATUS_FILE).read_text(encoding="utf-8"))
        except Exception:
            data = {}
    result = {}
    for name in names:
        result[name] = {**_default_status(), **data.get(name, {})}
    return result


def save_robot_arm_status(status: Dict[str, Dict]) -> None:
    Path(ROBOT_ARM_STATUS_FILE).write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")


def build_return_to_standby_route(
    current_location: str,
    vehicle_idx: int = 0,
    waypoint_rules: Optional[Dict[str, str]] = None,
) -> List["RouteStep"]:
    """current_location에서 이 차량(vehicle_idx)의 자기 대기 위치로 복귀하는 경로를 만듭니다
    (필수 경유지 규칙 반영). 이미 그 위치에 있으면 빈 리스트를 반환합니다."""
    home = vehicle_home_location(vehicle_idx)
    if current_location == home:
        return []
    if waypoint_rules is None:
        waypoint_rules = load_waypoint_rules()

    steps = [RouteStep(current_location, "출발")]  # 실제 출발 위치를 첫 단계로 명시 (build_route와 동일한 패턴)
    for name in expand_leg(current_location, home, waypoint_rules):
        action = "복귀" if name == home else "경유"
        steps.append(RouteStep(name, action))
    return steps


def describe_route_sentence(item_label: str, route: List["RouteStep"]) -> str:
    """경로를 "OO에서 출발해 XX를 거쳐 YY로 이동 완료" 형태의 문장 하나로 요약합니다."""
    if not route:
        return f"{item_label}: 이동할 경로가 없습니다."

    start = route[0].location
    end = route[-1].location
    via_names = [step.location for step in route[1:-1]]

    if via_names:
        via_text = "', '".join(via_names)
        return f"{item_label}: '{start}'에서 출발해 '{via_text}'를 거쳐 '{end}'로 이동 완료"
    return f"{item_label}: '{start}'에서 출발해 '{end}'로 이동 완료"


# -----------------------------------------------------------------------------
# 위치 데이터 로딩 (dual_view_calibrator.py / location_marker_tool.py 결과물 재사용)
# -----------------------------------------------------------------------------

def load_named_locations() -> Dict[str, Dict]:
    """
    두 종류의 저장 형식을 모두 지원합니다.
    - location_marks_verified.json: {"항구": {"cctv_pixel": [..], "map_meters": [..]}}
    - location_marks.json (구버전): {"항구": [px, py]}
    반환값은 항상 {"항구": {"cctv_pixel": [...], "map_meters": [...] (있으면)}} 형태로 통일합니다.
    """
    if Path(VERIFIED_LOCATIONS_FILE).exists():
        # 신버전 파일은 이미 원하는 형태(dict of dict)라서 그대로 반환
        raw = json.loads(Path(VERIFIED_LOCATIONS_FILE).read_text(encoding="utf-8"))
        return raw

    if Path(SIMPLE_LOCATIONS_FILE).exists():
        # 구버전 파일은 {"항구": [px, py]} 형태라서, 신버전과 같은 모양으로 감싸줌
        raw = json.loads(Path(SIMPLE_LOCATIONS_FILE).read_text(encoding="utf-8"))
        return {name: {"cctv_pixel": xy} for name, xy in raw.items()}

    return {}


# -----------------------------------------------------------------------------
# 화물 위치 등록부
# -----------------------------------------------------------------------------

CARGO_DETAILS_FILE = str(_APP_DIR / "cargo_details.json")
EXCEL_DATA_START_ROW = 6  # generate_cargo_template.py의 DATA_START_ROW와 일치해야 함


def load_cargo_registry() -> Dict[str, str]:
    if Path(CARGO_FILE).exists():
        return json.loads(Path(CARGO_FILE).read_text(encoding="utf-8"))
    return {}


def _sync_cargo_to_db(name: str, location: str, details: dict | None = None) -> None:
    """화물 한 건을 PostgreSQL cargos 테이블에 upsert 합니다.

    DB 연결이 불가능한 환경(개발 머신 등)에서는 조용히 무시하여 JSON
    파일 기반 동작이 깨지지 않도록 합니다.
    """
    try:
        detail = details or {}
        admin = InventoryAdminClient()
        admin.upsert_cargo(
            name=name,
            location=location,
            container_id=detail.get('컨테이너ID', ''),
            cargo_type=detail.get('화물종류', ''),
            note=detail.get('비고', ''),
            base_aruco_id=detail.get('기반ArUco', ''),
            floor=int(detail.get('층수', '1') or 1),
        )
    except Exception as exc:
        print(f'[DB 동기화 경고] {name} upsert 실패: {exc}')


def _delete_cargo_from_db(name: str) -> None:
    """화물 한 건을 PostgreSQL cargos 테이블에서 삭제합니다."""
    try:
        admin = InventoryAdminClient()
        admin.delete_cargo(name)
    except Exception as exc:
        print(f'[DB 동기화 경고] {name} delete 실패: {exc}')


def save_cargo_registry(registry: Dict[str, str]) -> None:
    Path(CARGO_FILE).write_text(json.dumps(registry, ensure_ascii=False, indent=2), encoding="utf-8")


def save_cargo_registry_and_db(
    registry: Dict[str, str], details: Dict[str, Dict[str, str]] | None = None
) -> None:
    """JSON 파일에 저장함과 동시에 PostgreSQL에 upsert 합니다."""
    Path(CARGO_FILE).write_text(json.dumps(registry, ensure_ascii=False, indent=2), encoding="utf-8")
    details = details if details is not None else load_cargo_details()
    for name, location in registry.items():
        _sync_cargo_to_db(name, location, details.get(name))


def load_cargo_details() -> Dict[str, Dict[str, str]]:
    if Path(CARGO_DETAILS_FILE).exists():
        return json.loads(Path(CARGO_DETAILS_FILE).read_text(encoding="utf-8"))
    return {}


def save_cargo_details(details: Dict[str, Dict[str, str]]) -> None:
    Path(CARGO_DETAILS_FILE).write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")


def save_cargo_details_and_db(
    registry: Dict[str, str], details: Dict[str, Dict[str, str]]
) -> None:
    """JSON 파일에 저장함과 동시에 PostgreSQL에 upsert 합니다."""
    Path(CARGO_DETAILS_FILE).write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")
    for name, detail in details.items():
        location = registry.get(name, '')
        if location:
            _sync_cargo_to_db(name, location, detail)


def bulk_import_cargo_from_excel(
    path: str,
    known_locations,
) -> Tuple[Dict[str, str], Dict[str, Dict[str, str]], List[str]]:
    """
    generate_cargo_template.py로 만든 양식을 읽어 (name->location, name->세부정보, 오류목록)을 반환합니다.
    known_locations에 없는 위치명이 있어도 등록은 하되, 오류 목록에 남겨서 사용자가 확인하게 합니다.
    """
    wb = load_workbook(path, data_only=True)  # data_only=True: 수식이 아니라 계산된 값만 읽음
    ws = wb["화물등록"] if "화물등록" in wb.sheetnames else wb.active  # 시트 이름이 바뀌었어도 안전하게 동작

    known_set = set(known_locations)  # in 검사를 빠르게 하려고 set으로 변환
    new_registry: Dict[str, str] = {}
    new_details: Dict[str, Dict[str, str]] = {}
    errors: List[str] = []

    # 헤더/예시 행을 건너뛰고 실제 데이터가 시작되는 6행부터 끝까지 한 줄씩 읽음
    for row_idx in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value          # A열: 화물명
        location = ws.cell(row=row_idx, column=2).value       # B열: 현재위치
        container_id = ws.cell(row=row_idx, column=3).value   # C열: 컨테이너/ArUco ID
        cargo_type = ws.cell(row=row_idx, column=4).value     # D열: 화물종류
        note = ws.cell(row=row_idx, column=5).value           # E열: 비고

        if name is None or str(name).strip() == "":
            continue  # 화물명이 비어있으면 빈 행으로 간주하고 건너뜀

        name = str(name).strip()
        location = str(location).strip() if location is not None else ""

        if not location:
            errors.append(f"{row_idx}행: '{name}'의 현재위치가 비어있어 건너뛰었습니다.")
            continue

        if known_set and location not in known_set:
            # 등록되지 않은 위치명이어도 일단 저장은 하되, 나중에 사용자가 확인할 수 있게 기록만 해둠
            errors.append(f"{row_idx}행: '{name}'의 위치 '{location}'는 등록된 위치 목록에 없습니다. (그대로 저장은 됨)")

        new_registry[name] = location
        new_details[name] = {
            "컨테이너ID": str(container_id).strip() if container_id else "",
            "화물종류": str(cargo_type).strip() if cargo_type else "",
            "비고": str(note).strip() if note else "",
        }

    return new_registry, new_details, errors


def plan_relocations_from_excel(
    path: str,
    old_registry: Dict[str, str],
    known_locations,
    vehicle_positions: Optional[List[str]] = None,
    waypoint_rules: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict], List[str], List[str], Dict[str, str], Dict[str, Dict[str, str]], List[str], int]:
    """
    엑셀(새 위치)과 기존 등록부(old_registry, 옛 위치)를 비교해서
    - jobs: 위치가 바뀐 화물들의 이동 경로 목록 [{"item":.., "route":[...], "container_id":..,
      "vehicle_index":.., "is_return":False/True, "label":..}] (맨 뒤에 차량 복귀 job도 포함)
    - unchanged: 위치 변경이 없는 화물명 목록
    - new_items: 이번에 처음 등록되는 화물명 목록 (이전 위치가 없어 경로 생성 대상 아님)
    - final_next_vehicle_index: 이번 배치에서 다 쓰고 난 뒤의 차량 순번(다음 배치에 이어서 씀)
    을 함께 반환합니다.

    차량 대수(NUM_VEHICLES)만큼은 대기장소에서 출발하고, 그 이후 화물부터는 직전에
    그 차량이 마지막으로 도착한 위치에서 출발하도록 순서대로 시뮬레이션합니다.
    배치 마지막에는 실제로 사용한 차량들을 대기장소로 복귀시키는 job도 추가됩니다.
    """
    if waypoint_rules is None:
        waypoint_rules = load_waypoint_rules()
    if vehicle_positions is None:
        vehicle_positions = [vehicle_home_location(i) for i in range(NUM_VEHICLES)]
    sim_positions = list(vehicle_positions)  # 원본을 훼손하지 않도록 복사본에서 시뮬레이션
    next_index = 0

    new_registry, new_details, errors = bulk_import_cargo_from_excel(path, known_locations)

    jobs: List[Dict] = []       # 위치가 바뀐 화물 -> 이동 경로가 필요한 것들
    unchanged: List[str] = []   # 예전과 위치가 같은 화물 -> 아무것도 안 해도 됨
    new_items: List[str] = []   # 등록부에 아예 없던 화물 -> "이동"이 아니라 "신규 등록" 대상
    used_vehicle_indices = set()

    for name, new_location in new_registry.items():
        old_location = old_registry.get(name)  # 기존 등록부에서 이 화물의 예전 위치 조회

        if old_location is None:
            new_items.append(name)  # 예전 기록이 없으니 이동이 아니라 신규 등록으로 처리
            continue

        if old_location == new_location:
            unchanged.append(name)  # 위치가 그대로면 재배차 대상 아님
            continue

        vehicle_idx = next_index % len(sim_positions)
        start_location = sim_positions[vehicle_idx]
        next_index += 1

        # 위치가 실제로 바뀐 경우에만 [차량 현재위치 -> 예전위치(픽업) -> 새위치(하역)] 경로를 생성
        # build_route는 원래 "화물등록부에서 위치를 조회"하는 함수인데, 여기서는 이미 old_location을
        # 알고 있으므로 {name: old_location} 라는 임시 1개짜리 등록부를 만들어 그대로 재사용합니다.
        route = build_route(name, new_location, {name: old_location},
                            standby_location=start_location, waypoint_rules=waypoint_rules)
        
        is_crane_only = any(step.action == "크레인 전용 이동" for step in route)
        if is_crane_only:
            next_index -= 1 # 배차 취소 (차량 이동 없음)
            jobs.append({
                "item": name,
                "route": route,
                "container_id": new_details.get(name, {}).get("컨테이너ID", ""),
                "vehicle_index": -1,
                "is_return": False,
                "label": name,
            })
            continue

        agv_final_loc = route[-1].location
        for step in reversed(route):
            if step.action not in ["크레인 최종 이동", "크레인 전용 이동"]:
                agv_final_loc = step.location
                break

        sim_positions[vehicle_idx] = agv_final_loc
        used_vehicle_indices.add(vehicle_idx)

        jobs.append({
            "item": name,
            "route": route,
            "container_id": new_details.get(name, {}).get("컨테이너ID", ""),
            "vehicle_index": vehicle_idx,
            "is_return": False,
            "label": name,
        })

    # 배치 마지막에 실제로 쓰인 차량들을 각자의 대기 위치로 복귀시키는 job을 큐 맨 뒤에 추가
    for vehicle_idx in sorted(used_vehicle_indices):
        current = sim_positions[vehicle_idx]
        if "창고 하역장" in current:
            continue  # 창고 하역장에 정차 유지

        return_route = build_return_to_standby_route(current, vehicle_idx, waypoint_rules)
        if not return_route:
            continue  # 이미 자기 대기 위치에 있음
        jobs.append({
            "item": None,
            "route": return_route,
            "container_id": "",
            "vehicle_index": vehicle_idx,
            "is_return": True,
            "label": f"차량 {vehicle_idx + 1} 복귀",
        })
        sim_positions[vehicle_idx] = vehicle_home_location(vehicle_idx)

    return jobs, unchanged, new_items, new_registry, new_details, errors, next_index % len(sim_positions)

    return jobs, unchanged, new_items, new_registry, new_details, errors


# -----------------------------------------------------------------------------
# 경로 계획
# -----------------------------------------------------------------------------

class RouteStep:
    def __init__(self, location: str, action: str):
        self.location = location  # 위치 이름
        self.action = action      # "출발" / "적재" / "하역"

    def __repr__(self):
        return f"{self.location}({self.action})"


def build_route(
    item_name: str,
    destination: str,
    cargo_registry: Dict[str, str],
    standby_location: str = STANDBY_LOCATION,
    waypoint_rules: Optional[Dict[str, str]] = None,
) -> List[RouteStep]:
    """
    화물 등록부에서 item_name의 현재 위치(픽업 지점)를 찾아
    [대기장소 -> (필수 경유지 있으면 경유) -> 픽업 -> (필수 경유지 있으면 경유) -> 목적지]
    순서의 경로를 만듭니다. 화물이 이미 목적지에 있으면 픽업 단계 없이 안내만 합니다.
    """
    if item_name not in cargo_registry:
        raise ValueError(f"'{item_name}'의 현재 위치를 알 수 없습니다. 먼저 화물 위치를 등록해주세요.")

    if waypoint_rules is None:
        waypoint_rules = load_waypoint_rules()

    pickup = cargo_registry[item_name]  # 이 화물이 지금 실제로 놓여있는 위치

    if pickup == destination:
        # 이미 목적지에 있으면 이동할 필요가 없으니 안내용 1단계짜리 경로만 반환
        return [RouteStep(pickup, "이미 목적지에 있음")]

    is_pickup_warehouse = "창고" in pickup and "하역장" not in pickup
    is_pickup_port = "항구" in pickup and "하역장" not in pickup
    is_dest_warehouse = "창고" in destination and "하역장" not in destination
    is_dest_port = "항구" in destination and "하역장" not in destination

    # 1) 창고 -> 창고 (크레인 전용 이동)
    if is_pickup_warehouse and is_dest_warehouse:
        return [
            RouteStep(pickup, "크레인 상차"),
            RouteStep(destination, "크레인 전용 이동")
        ]

    agv_pickup = pickup
    agv_destination = destination

    # 2) 창고 -> 항구
    if is_pickup_warehouse and is_dest_port:
        agv_pickup = "창고 하역장"
        agv_destination = "항구 하역장"
    # 3) 항구 -> 창고
    elif is_pickup_port and is_dest_warehouse:
        agv_pickup = "항구 하역장"
        agv_destination = "창고 하역장"

    steps = [RouteStep(standby_location, "출발")]  # 차량은 항상 대기장소에서 출발
    current = standby_location  # 지금까지 경로상 마지막으로 도달한 위치 (다음 구간 계산의 출발점)

    if agv_pickup != standby_location:
        # [대기장소 -> 픽업] 구간에 필수 경유지가 있으면 expand_leg가 자동으로 끼워줌
        # 예: expand_leg("대기장소", "창고 하역장", rules) -> ["회차지점B", "창고 하역장"]
        for name in expand_leg(current, agv_pickup, waypoint_rules):
            action = "적재 (크레인 연계)" if name == agv_pickup else "경유"
            steps.append(RouteStep(name, action))
        current = agv_pickup  # 픽업까지 왔으니, 다음 구간은 여기서부터 계산

    # [픽업 -> 목적지] 구간도 마찬가지로 필수 경유지를 자동으로 끼워 넣음
    for name in expand_leg(current, agv_destination, waypoint_rules):
        action = "하역 (크레인 연계)" if name == agv_destination else "경유"
        steps.append(RouteStep(name, action))

    # AGV 이동 후, 화물이 최종 목적지로 크레인을 통해 이동함을 명시
    if agv_destination != destination:
        steps.append(RouteStep(destination, "크레인 최종 이동"))

    return steps


def location_coord_text(name: str, locations: Dict[str, Dict]) -> str:
    entry = locations.get(name)
    if not entry:
        return f"{name} (좌표 미등록)"
    if "map_meters" in entry:
        mx, my = entry["map_meters"]
        return f"{name} (map {mx:.2f}, {my:.2f} m)"
    if "cctv_pixel" in entry:
        px, py = entry["cctv_pixel"]
        return f"{name} (CCTV px {px:.0f}, {py:.0f})"
    return name


# -----------------------------------------------------------------------------
# GUI
# -----------------------------------------------------------------------------

def write_cargo_excel(cargo_registry: Dict[str, str], cargo_details: Dict[str, Dict[str, str]], output_path: str) -> None:
    """현재 화물 위치/세부정보를 엑셀로 내보냅니다. generate_cargo_template.py와 같은
    형식(화물등록 시트, 6행부터 데이터)으로 저장하기 때문에, 이 파일을 그대로
    "엑셀로 일괄 등록" 또는 "재배치 엑셀 불러오기"로 다시 불러올 수 있습니다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "화물등록"

    ws.merge_cells("A1:E1")
    ws["A1"] = "화물 위치 내보내기 (현재 상태 스냅샷)"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"내보낸 시각: {time.strftime('%Y-%m-%d %H:%M:%S')} - "
        "이 파일은 '엑셀로 일괄 등록' 또는 '재배치 엑셀 불러오기'로 다시 불러올 수 있습니다."
    )
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="808080")

    headers = ["화물명", "현재위치", "컨테이너/ArUco ID", "화물종류", "비고"]
    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=title)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [16, 16, 18, 14, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    row_idx = EXCEL_DATA_START_ROW
    for name, location in cargo_registry.items():
        detail = cargo_details.get(name, {})
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=location)
        ws.cell(row=row_idx, column=3, value=detail.get("컨테이너ID", ""))
        ws.cell(row=row_idx, column=4, value=detail.get("화물종류", ""))
        ws.cell(row=row_idx, column=5, value=detail.get("비고", ""))
        row_idx += 1

    wb.save(output_path)


class CargoDispatchTool(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        inventory_client = kwargs.pop("inventory_client", None)
        inventory_admin_client = kwargs.pop("inventory_admin_client", None)
        super().__init__(master, fg_color="transparent", **kwargs)

        self.font_title = ctk.CTkFont(family="Malgun Gothic", size=20, weight="bold")
        self.font_subtitle = ctk.CTkFont(family="Malgun Gothic", size=14, weight="bold")
        self.font_body = ctk.CTkFont(family="Malgun Gothic", size=12)

        self.locations = load_named_locations()       # 등록된 위치 이름과 좌표들
        self.cargo_registry = load_cargo_registry()   # {화물명: 현재위치} - 실제 배차 로직이 참조하는 핵심 데이터
        self.cargo_details = load_cargo_details()     # {화물명: {컨테이너ID, 화물종류, 비고}} - 부가정보
        self.rules = load_waypoint_rules()            # 필수 경유지 규칙 (build_route에 전달)
        self.vehicle_positions, self.next_vehicle_index = load_vehicle_state()  # 차량별 현재 위치

        # PostgreSQL is authoritative for the inventory list shown on this page.
        # Queries run outside the Tk thread so a DB timeout cannot freeze the UI.
        self.inventory_client = inventory_client or InventoryClient()
        self.inventory_admin_client = (
            inventory_admin_client
            or InventoryAdminClient(self.inventory_client)
        )
        self._inventory_records = None
        self._inventory_snapshot_id = ""
        self._inventory_last_success = ""
        self._inventory_error = ""
        self._inventory_poll_in_flight = False
        self._inventory_reset_in_flight = False
        self._inventory_delete_in_flight = False
        self._inventory_stopped = False
        self._inventory_after_id = None
        self._inventory_results = queue.Queue()
        self._inventory_reset_results = queue.Queue()
        self._inventory_delete_results = queue.Queue()
        self._inventory_refresh_ms = _inventory_refresh_interval_ms()

        # 배차 실행 큐 상태 - 자연어 단건 명령이든 엑셀 다건 재배치든 이 큐 하나로 통일해서 처리
        self.dispatch_queue: List[Dict] = []  # [{"item": 화물명, "route": [RouteStep, ...]}, ...]
        self.active_job_index: int = 0        # 큐에서 지금 처리 중인 화물(job)의 인덱스
        self.active_step_index: int = 0       # 그 화물의 경로 중 지금 처리 중인 단계(step)의 인덱스

        self._build_ui()
        self._refresh_cargo_list()
        self._request_inventory_refresh()

    # ------------------------------------------------------------------
    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=5)
        self.grid_columnconfigure(1, weight=7)
        self.grid_rowconfigure(1, weight=1)
        self.configure(fg_color="#242424")

        # Header
        header_frame = ctk.CTkFrame(self, height=56, fg_color="#2b2b2b", corner_radius=0, border_width=1, border_color="#1a1a1a")
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
        header_frame.pack_propagate(False)
        ctk.CTkLabel(header_frame, text="🚚 화물 위치 추적 및 자연어 배차", font=self.font_title, text_color="#dce4ee").pack(side="left", padx=24, pady=16)

        # ---- Left Panel: Cargo Registration ----
        left = ctk.CTkFrame(self, fg_color="transparent")
        left.grid(row=1, column=0, sticky="nsew", padx=(24, 12), pady=24)
        left.grid_columnconfigure(0, weight=1)

        # 1) Registration Form Card
        reg_card = ctk.CTkFrame(left, fg_color="#2b2b2b", corner_radius=6, border_width=1, border_color="#1a1a1a")
        reg_card.grid(row=0, column=0, sticky="ew", pady=(0, 24))
        
        ctk.CTkLabel(reg_card, text="화물 위치 등록", font=self.font_subtitle, text_color="#dce4ee").pack(anchor="w", padx=20, pady=(20, 16))

        form = ctk.CTkFrame(reg_card, fg_color="transparent")
        form.pack(fill="x", padx=20, pady=(0, 20))
        form.grid_columnconfigure(0, weight=1)

        entry_kwargs = {"fg_color": "#343638", "border_color": "#565b5e", "text_color": "#dce4ee", "corner_radius": 6, "height": 36}
        self.container_id_var = ctk.StringVar(value=MANUAL_CONTAINER_IDS[0])
        ctk.CTkLabel(
            form, text="컨테이너 ID", font=self.font_body,
            text_color="#9e9e9e",
        ).grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.container_id_menu = ctk.CTkOptionMenu(
            form,
            variable=self.container_id_var,
            values=list(MANUAL_CONTAINER_IDS),
            fg_color="#343638",
            button_color="#565b5e",
            button_hover_color="#2e86c1",
            corner_radius=6,
            height=36,
        )
        self.container_id_menu.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        self.location_option_var = ctk.StringVar()
        ctk.CTkLabel(
            form, text="표준 위치", font=self.font_body,
            text_color="#9e9e9e",
        ).grid(row=2, column=0, sticky="w", pady=(0, 4))
        location_names = list(MANUAL_INVENTORY_LOCATIONS)
        self.location_option_var.set(location_names[0])
        self.location_menu = ctk.CTkOptionMenu(form, variable=self.location_option_var, values=location_names,
                                              fg_color="#343638", button_color="#565b5e", button_hover_color="#2e86c1", corner_radius=6, height=36)
        self.location_menu.grid(row=3, column=0, sticky="ew", pady=(0, 12))

        self.floor_var = ctk.StringVar(value='1')
        ctk.CTkLabel(
            form, text="층수", font=self.font_body,
            text_color="#9e9e9e",
        ).grid(row=4, column=0, sticky="w", pady=(0, 4))
        self.floor_menu = ctk.CTkOptionMenu(
            form,
            variable=self.floor_var,
            values=list(MANUAL_FLOORS),
            fg_color="#343638",
            button_color="#565b5e",
            button_hover_color="#2e86c1",
            corner_radius=6,
            height=36,
        )
        self.floor_menu.grid(row=5, column=0, sticky="ew", pady=(0, 12))

        btn_kwargs = {"corner_radius": 6, "height": 36, "font": self.font_body}
        
        ctk.CTkButton(form, text="DB 위치 등록 / 갱신", fg_color="#2e86c1", hover_color="#21618c", command=self.register_cargo_location, **btn_kwargs).grid(row=6, column=0, sticky="ew", pady=4)
        
        # Detail Banner
        self.detail_banner = ctk.CTkFrame(form, fg_color="#343638", corner_radius=6, border_width=1, border_color="#2e86c1")
        ctk.CTkLabel(self.detail_banner, text="📋 추가 정보 입력 (선택사항)", font=self.font_subtitle, text_color="#2e86c1").grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=10)
        
        self.cargo_type_var = ctk.StringVar()
        ctk.CTkLabel(self.detail_banner, text="화물 종류:", font=self.font_body).grid(row=1, column=0, sticky="w", padx=10, pady=2)
        ctk.CTkEntry(self.detail_banner, textvariable=self.cargo_type_var, placeholder_text="예: 컨테이너, 팔레트", width=160, **entry_kwargs).grid(row=1, column=1, sticky="ew", padx=(0, 10), pady=2)
        
        self.cargo_note_var = ctk.StringVar()
        ctk.CTkLabel(self.detail_banner, text="비고:", font=self.font_body).grid(row=2, column=0, sticky="w", padx=10, pady=2)
        ctk.CTkEntry(self.detail_banner, textvariable=self.cargo_note_var, placeholder_text="메모", width=160, **entry_kwargs).grid(row=2, column=1, sticky="ew", padx=(0, 10), pady=2)
        
        banner_btn_row = ctk.CTkFrame(self.detail_banner, fg_color="transparent")
        banner_btn_row.grid(row=3, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
        ctk.CTkButton(banner_btn_row, text="✅ 저장", width=80, fg_color="#27ae60", hover_color="#1e8449", command=self._confirm_register, **btn_kwargs).pack(side="left", padx=(0, 6))
        ctk.CTkButton(banner_btn_row, text="건너뛰기", width=80, fg_color="#565b5e", hover_color="#333333", command=self._skip_details_register, **btn_kwargs).pack(side="left", padx=(0, 6))
        ctk.CTkButton(banner_btn_row, text="취소", width=60, fg_color="#c0392b", hover_color="#922b21", command=self._cancel_register, **btn_kwargs).pack(side="left")

        self._pending_cargo_name: Optional[str] = None
        self._pending_container_id: Optional[str] = None
        self._pending_cargo_location: Optional[str] = None
        self._pending_cargo_floor: Optional[int] = None
        self._pending_base_aruco_id: Optional[str] = None

        ctk.CTkButton(form, text="📄 엑셀로 일괄 등록", fg_color="#27ae60", hover_color="#1e8449", command=self.bulk_import_from_excel, **btn_kwargs).grid(row=8, column=0, sticky="ew", pady=4)
        ctk.CTkButton(form, text="🧾 빈 엑셀 양식 생성", fg_color="#343638", hover_color="#333333", border_width=1, border_color="#565b5e", text_color="#dce4ee", command=self.generate_template, **btn_kwargs).grid(row=9, column=0, sticky="ew", pady=4)
        ctk.CTkButton(form, text="📤 현재 화물정보 엑셀로 내보내기", fg_color="#f39c12", hover_color="#d68910", command=self.export_cargo_to_excel, **btn_kwargs).grid(row=10, column=0, sticky="ew", pady=4)

        # ---- Right Panel: Live PostgreSQL inventory ----
        right = ctk.CTkFrame(self, fg_color="transparent")
        right.grid(row=1, column=1, sticky="nsew", padx=(12, 24), pady=24)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)

        list_card = ctk.CTkFrame(right, fg_color="#2b2b2b", corner_radius=6, border_width=1, border_color="#1a1a1a")
        list_card.grid(row=0, column=0, sticky="nsew")
        list_card.grid_columnconfigure(0, weight=1)
        list_card.grid_rowconfigure(2, weight=1)

        list_header = ctk.CTkFrame(list_card, fg_color="#2b2b2b", height=52, corner_radius=6)
        list_header.grid(row=0, column=0, sticky="ew")
        list_header.pack_propagate(False)
        ctk.CTkLabel(list_header, text="DB 실시간 화물 목록", font=self.font_subtitle, text_color="#dce4ee").pack(side="left", padx=16, pady=16)
        ctk.CTkButton(list_header, text="🔄 DB 조회", width=80, fg_color="transparent", text_color="#2e86c1", hover_color="#333333", command=self._request_inventory_refresh).pack(side="right", padx=16, pady=16)
        ctk.CTkButton(
            list_header,
            text="🗣️ 명령",
            width=80,
            fg_color="#92ccff",
            text_color="#003351",
            hover_color="#cce5ff",
            command=self.open_command_popup,
        ).pack(side="right", padx=(0, 4), pady=10)
        self.inventory_reset_button = ctk.CTkButton(
            list_header,
            text="DB 초기화",
            width=76,
            fg_color="#c0392b",
            hover_color="#922b21",
            text_color="white",
            command=self._confirm_inventory_reset,
        )
        self.inventory_reset_button.pack(side="right", padx=(0, 4), pady=16)

        ctk.CTkFrame(list_card, height=1, fg_color="#1a1a1a").grid(row=0, column=0, sticky="sew")

        self.inventory_status_label = ctk.CTkLabel(
            list_card,
            text="PostgreSQL 조회 대기 중",
            font=self.font_body,
            text_color="#f0ad4e",
            anchor="w",
            justify="left",
            wraplength=520,
        )
        self.inventory_status_label.grid(row=1, column=0, sticky="ew", padx=16, pady=(10, 0))

        self.cargo_rows_container = ctk.CTkScrollableFrame(list_card, fg_color="#242424", border_width=1, border_color="#1a1a1a", corner_radius=6)
        self.cargo_rows_container.grid(row=2, column=0, sticky="nsew", padx=16, pady=16)
        self.cargo_rows_container.grid_columnconfigure(0, weight=1)

    def open_command_popup(self) -> None:
        from command_center import open_command_popup
        # 팝업에서 화물 위치를 바꾸면, 지금 이 화면(이미 메모리에 옛 데이터를 들고 있음)이
        # 자동으로 최신 상태로 갱신되도록 콜백을 넘겨줍니다.
        open_command_popup(self, on_cargo_updated=self.refresh_from_disk)

    def refresh_from_disk(self) -> None:
        """위치/화물 데이터를 디스크에서 다시 읽어와 화면을 최신 상태로 갱신합니다.
        팝업에서 명령을 실행한 직후 자동으로 호출되며, "🔄 새로고침" 버튼으로 수동으로도
        호출할 수 있습니다 (예: 다른 창에서 엑셀로 데이터를 바꾼 경우 등)."""
        self.locations = load_named_locations()
        self.cargo_registry = load_cargo_registry()
        self.cargo_details = load_cargo_details()
        self._refresh_cargo_list()
        self._request_inventory_refresh()

    def _request_inventory_refresh(self) -> None:
        """Start one non-blocking PostgreSQL inventory query."""
        if (
            self._inventory_stopped
            or self._inventory_poll_in_flight
            or self._inventory_reset_in_flight
            or self._inventory_delete_in_flight
        ):
            return
        pending_after_id = self._inventory_after_id
        self._inventory_after_id = None
        if pending_after_id is not None:
            try:
                self.after_cancel(pending_after_id)
            except Exception:
                pass
        self._inventory_poll_in_flight = True
        if hasattr(self, "inventory_status_label"):
            self.inventory_status_label.configure(
                text="PostgreSQL 실시간 조회 중…",
                text_color="#f0ad4e",
            )

        def fetch() -> None:
            try:
                self._inventory_results.put((True, self.inventory_client.fetch_snapshot()))
            except Exception as exc:
                self._inventory_results.put((False, exc))

        threading.Thread(
            target=fetch,
            name="cargo-inventory-ui",
            daemon=True,
        ).start()
        self._schedule_inventory_result_check(100)

    def _confirm_inventory_reset(self) -> None:
        """Require operator confirmation before deleting all inventory rows."""
        if self._inventory_reset_in_flight or self._inventory_delete_in_flight:
            return
        if self._inventory_poll_in_flight:
            messagebox.showinfo(
                "DB 조회 중",
                "현재 DB 조회가 끝난 뒤 초기화 버튼을 다시 눌러주세요.",
            )
            return
        confirmed = messagebox.askyesno(
            "DB 전체 초기화",
            "cargos의 현재 화물 정보와 cargo_movements의 전체 이동 이력을 모두 삭제합니다.\n\n"
            "자율 관제모드를 먼저 중지해야 하며, 삭제된 정보는 되돌릴 수 없습니다.\n"
            "계속하시겠습니까?",
            icon="warning",
        )
        if not confirmed:
            return

        pending_after_id = self._inventory_after_id
        self._inventory_after_id = None
        if pending_after_id is not None:
            try:
                self.after_cancel(pending_after_id)
            except Exception:
                pass
        self._inventory_reset_in_flight = True
        self.inventory_reset_button.configure(state="disabled", text="초기화 중…")
        self.inventory_status_label.configure(
            text="PostgreSQL 화물 정보와 이동 이력을 초기화하는 중…",
            text_color="#f0ad4e",
        )

        def reset() -> None:
            try:
                self.inventory_admin_client.clear_inventory()
                self._inventory_reset_results.put((True, None))
            except Exception as exc:
                self._inventory_reset_results.put((False, exc))

        threading.Thread(
            target=reset,
            name="cargo-inventory-reset-ui",
            daemon=True,
        ).start()
        self._schedule_inventory_reset_check(100)

    def _schedule_inventory_reset_check(self, delay_ms: int) -> None:
        if self._inventory_stopped:
            return
        try:
            self._inventory_after_id = self.after(
                delay_ms,
                self._consume_inventory_reset_result,
            )
        except Exception:
            self._inventory_stopped = True

    def _confirm_inventory_delete(self, cargo) -> None:
        """Delete one live PostgreSQL cargo row after operator confirmation."""
        if self._inventory_delete_in_flight or self._inventory_reset_in_flight:
            return
        if self._inventory_poll_in_flight:
            messagebox.showinfo(
                "DB 조회 중",
                "현재 DB 조회가 끝난 뒤 삭제 버튼을 다시 눌러주세요.",
            )
            return

        name = str(cargo.name)
        container_id = str(cargo.container_id or '-')
        confirmed = messagebox.askyesno(
            "DB 화물 개별 삭제",
            f"'{name}' (컨테이너 ID {container_id})을 DB에서 삭제할까요?\n\n"
            "현재 위치 정보만 삭제하며 기존 이동 이력은 유지됩니다.\n"
            "자율 관제 중이라면 먼저 자율 관제모드를 중지해주세요.",
            icon="warning",
        )
        if not confirmed:
            return

        pending_after_id = self._inventory_after_id
        self._inventory_after_id = None
        if pending_after_id is not None:
            try:
                self.after_cancel(pending_after_id)
            except Exception:
                pass
        self._inventory_delete_in_flight = True
        self.inventory_reset_button.configure(state="disabled")
        self.inventory_status_label.configure(
            text=f"PostgreSQL에서 {name}을 삭제하는 중…",
            text_color="#f0ad4e",
        )
        self._refresh_cargo_list()

        def delete_one() -> None:
            try:
                self.inventory_admin_client.delete_cargo(name)
                self._inventory_delete_results.put((True, name))
            except Exception as exc:
                self._inventory_delete_results.put((False, (name, exc)))

        threading.Thread(
            target=delete_one,
            name="cargo-inventory-delete-ui",
            daemon=True,
        ).start()
        self._schedule_inventory_delete_check(100)

    def _schedule_inventory_delete_check(self, delay_ms: int) -> None:
        if self._inventory_stopped:
            return
        try:
            self._inventory_after_id = self.after(
                delay_ms,
                self._consume_inventory_delete_result,
            )
        except Exception:
            self._inventory_stopped = True

    def _consume_inventory_delete_result(self) -> None:
        if self._inventory_stopped:
            return
        try:
            success, result = self._inventory_delete_results.get_nowait()
        except queue.Empty:
            self._schedule_inventory_delete_check(100)
            return

        self._inventory_after_id = None
        self._inventory_delete_in_flight = False
        self.inventory_reset_button.configure(state="normal")
        if success:
            name = str(result)
            self._inventory_records = tuple(
                cargo for cargo in (self._inventory_records or ())
                if cargo.name != name
            )
            self._inventory_snapshot_id = ""
            self._inventory_last_success = time.strftime("%H:%M:%S")
            self._inventory_error = ""
            self.cargo_registry.pop(name, None)
            self.cargo_details.pop(name, None)
            save_cargo_registry(self.cargo_registry)
            save_cargo_details(self.cargo_details)
            self.inventory_status_label.configure(
                text=f"● {name} 개별 삭제 완료 · 최신 DB를 다시 조회합니다.",
                text_color="#61de8a",
            )
            self._refresh_cargo_list()
            self._request_inventory_refresh()
            return

        name, error = result
        self._inventory_error = " ".join(str(error).split())[:220]
        self.inventory_status_label.configure(
            text=f"● {name} 삭제 실패 · {self._inventory_error}",
            text_color="#ff6b6b",
        )
        messagebox.showerror(
            "DB 화물 삭제 실패",
            f"PostgreSQL에서 {name}을 삭제하지 못했습니다.\n\n"
            f"{self._inventory_error}",
        )
        try:
            self._inventory_after_id = self.after(
                self._inventory_refresh_ms,
                self._request_inventory_refresh,
            )
        except Exception:
            self._inventory_stopped = True

    def _consume_inventory_reset_result(self) -> None:
        if self._inventory_stopped:
            return
        try:
            success, result = self._inventory_reset_results.get_nowait()
        except queue.Empty:
            self._schedule_inventory_reset_check(100)
            return

        self._inventory_after_id = None
        self._inventory_reset_in_flight = False
        self.inventory_reset_button.configure(state="normal", text="DB 초기화")
        if success:
            self._inventory_records = tuple()
            self._inventory_snapshot_id = ""
            self._inventory_last_success = time.strftime("%H:%M:%S")
            self._inventory_error = ""
            self.inventory_status_label.configure(
                text="● DB 초기화 완료 · 최신 상태를 다시 조회합니다.",
                text_color="#61de8a",
            )
            self._refresh_cargo_list()
            self._request_inventory_refresh()
            return

        self._inventory_error = " ".join(str(result).split())[:220]
        self.inventory_status_label.configure(
            text=f"● DB 초기화 실패 · {self._inventory_error}",
            text_color="#ff6b6b",
        )
        messagebox.showerror(
            "DB 초기화 실패",
            f"PostgreSQL을 초기화하지 못했습니다.\n\n{self._inventory_error}",
        )
        try:
            self._inventory_after_id = self.after(
                self._inventory_refresh_ms,
                self._request_inventory_refresh,
            )
        except Exception:
            self._inventory_stopped = True

    def _schedule_inventory_result_check(self, delay_ms: int) -> None:
        if self._inventory_stopped:
            return
        try:
            self._inventory_after_id = self.after(delay_ms, self._consume_inventory_result)
        except Exception:
            self._inventory_stopped = True

    def _consume_inventory_result(self) -> None:
        if self._inventory_stopped:
            return
        try:
            success, result = self._inventory_results.get_nowait()
        except queue.Empty:
            self._schedule_inventory_result_check(100)
            return

        self._inventory_poll_in_flight = False
        self._inventory_after_id = None
        refresh_rows = False
        if success:
            refresh_rows = (
                result.snapshot_id != self._inventory_snapshot_id
                or self._inventory_records is None
                or bool(self._inventory_error)
            )
            self._inventory_records = tuple(result.cargos)
            self._inventory_snapshot_id = result.snapshot_id
            self._inventory_last_success = time.strftime("%H:%M:%S")
            self._inventory_error = ""
            self.inventory_status_label.configure(
                text=(
                    f"● DB 실시간 연결 · {len(self._inventory_records)}건 · "
                    f"{self._inventory_last_success} · 스냅샷 {self._inventory_snapshot_id}"
                ),
                text_color="#61de8a",
            )
        else:
            refresh_rows = not self._inventory_error
            self._inventory_error = " ".join(str(result).split())[:220]
            if self._inventory_last_success:
                status = (
                    f"● DB 갱신 실패 · 아래 정보는 {self._inventory_last_success} 기준이며 최신이 아닙니다 · "
                    f"{self._inventory_error}"
                )
            else:
                status = f"● DB 연결 실패 · {self._inventory_error}"
            self.inventory_status_label.configure(text=status, text_color="#ff6b6b")

        # Preserve the operator's scroll position when the DB content is unchanged.
        if refresh_rows:
            self._refresh_cargo_list()
        try:
            self._inventory_after_id = self.after(
                self._inventory_refresh_ms,
                self._request_inventory_refresh,
            )
        except Exception:
            self._inventory_stopped = True

    def stop(self) -> None:
        """Stop dashboard polling when the user leaves this page."""
        self._inventory_stopped = True
        if self._inventory_after_id is not None:
            try:
                self.after_cancel(self._inventory_after_id)
            except Exception:
                pass
            self._inventory_after_id = None

    def destroy(self) -> None:
        self.stop()
        super().destroy()

    # ------------------------------------------------------------------
    # 화물 위치 등록
    # ------------------------------------------------------------------
    def register_cargo_location(self) -> None:
        """Validate ID/location/floor and open the optional detail editor."""
        container_id = self.container_id_var.get().strip()
        location = self.location_option_var.get().strip()
        try:
            floor = int(self.floor_var.get().strip())
            base_aruco_id = validate_manual_inventory_position(
                container_id,
                location,
                floor,
                self._inventory_records,
            )
        except ValueError as exc:
            messagebox.showerror("DB 위치 입력 오류", str(exc))
            return

        existing_cargo = next((
            cargo for cargo in (self._inventory_records or ())
            if str(cargo.container_id or '') == container_id
        ), None)
        name = (
            existing_cargo.name
            if existing_cargo is not None
            else f'컨테이너_C{container_id}'
        )
        self._pending_cargo_name = name
        self._pending_container_id = container_id
        self._pending_cargo_location = location
        self._pending_cargo_floor = floor
        self._pending_base_aruco_id = base_aruco_id

        existing = self.cargo_details.get(name, {})
        self.cargo_type_var.set(
            existing_cargo.cargo_type
            if existing_cargo is not None
            else existing.get("화물종류", "미분류")
        )
        self.cargo_note_var.set(
            existing_cargo.note
            if existing_cargo is not None
            else existing.get("비고", "운영자 수동 위치 등록")
        )

        self.detail_banner.grid(
            row=7, column=0, columnspan=2, sticky="ew", pady=(4, 4)
        )

    def _confirm_register(self) -> None:
        """배너의 '저장' 버튼 - 추가 정보를 포함해서 화물을 등록합니다."""
        details = {
            "컨테이너ID": self._pending_container_id or '',
            "화물종류": self.cargo_type_var.get().strip(),
            "비고": self.cargo_note_var.get().strip(),
            "기반ArUco": self._pending_base_aruco_id or '',
            "층수": str(self._pending_cargo_floor or 1),
        }
        self._save_pending_cargo(details)

    def _skip_details_register(self) -> None:
        """배너의 '건너뛰기' 버튼 - 추가 정보 없이 이름/위치만 저장합니다."""
        name = self._pending_cargo_name or ''
        existing = dict(self.cargo_details.get(name, {}))
        existing.update({
            "컨테이너ID": self._pending_container_id or '',
            "기반ArUco": self._pending_base_aruco_id or '',
            "층수": str(self._pending_cargo_floor or 1),
        })
        self._save_pending_cargo(existing)

    def _save_pending_cargo(self, details: dict) -> None:
        """Commit an operator-approved manual position to PostgreSQL."""
        name = self._pending_cargo_name
        location = self._pending_cargo_location
        container_id = self._pending_container_id
        floor = self._pending_cargo_floor
        if not name or not location or container_id is None or floor is None:
            return
        try:
            self.inventory_admin_client.upsert_cargo(
                name=name,
                location=location,
                container_id=container_id,
                cargo_type=details.get('화물종류', '') or '미분류',
                note=details.get('비고', '') or '운영자 수동 위치 등록',
                base_aruco_id=details.get('기반ArUco', ''),
                floor=floor,
            )
        except Exception as exc:
            messagebox.showerror(
                'DB 위치 저장 실패',
                f'PostgreSQL에 컨테이너 위치를 저장하지 못했습니다.\n\n{exc}',
            )
            return

        self.cargo_registry[name] = location
        self.cargo_details[name] = dict(details)
        save_cargo_registry(self.cargo_registry)
        save_cargo_details(self.cargo_details)
        self._hide_detail_banner()
        self._request_inventory_refresh()

    def _cancel_register(self) -> None:
        """배너의 '취소' 버튼 - 등록을 취소하고 배너를 숨깁니다."""
        self._hide_detail_banner()

    def _hide_detail_banner(self) -> None:
        """배너를 숨기고 입력 필드/임시 변수를 초기화합니다."""
        self.detail_banner.grid_forget()
        self._pending_cargo_name = None
        self._pending_container_id = None
        self._pending_cargo_location = None
        self._pending_cargo_floor = None
        self._pending_base_aruco_id = None
        self.cargo_type_var.set("")
        self.cargo_note_var.set("")

    def delete_cargo(self, name: str) -> None:
        """화물 목록의 각 행에 있는 "삭제" 버튼에서 호출됩니다. name이 이미 정해져 있으니
        이름을 따로 입력받지 않고 확인 창만 띄우고 바로 지웁니다."""
        if name not in self.cargo_registry:
            return
        if not messagebox.askyesno("화물 삭제", f"'{name}' 화물을 삭제할까요?"):
            return

        del self.cargo_registry[name]
        self.cargo_details.pop(name, None)  # 세부정보도 같이 있으면 제거, 없으면 조용히 무시
        save_cargo_registry(self.cargo_registry)
        save_cargo_details(self.cargo_details)
        _delete_cargo_from_db(name)
        self._refresh_cargo_list()

    def generate_template(self) -> None:
        """generate_cargo_template.py의 build_template()를 호출해서 빈 엑셀 양식을 만들어줍니다."""
        try:
            build_cargo_excel_template()
        except Exception as exc:
            messagebox.showerror("생성 실패", f"엑셀 양식 생성 중 오류가 발생했습니다:\n{exc}")
            return
        messagebox.showinfo("생성 완료", "cargo_template.xlsx 파일이 현재 폴더에 생성되었습니다.")

    def export_cargo_to_excel(self) -> None:
        """지금 화면에 있는 화물 위치/세부정보 전체를 엑셀 파일로 내보냅니다."""
        default_name = f"cargo_export_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="화물 정보 내보내기",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not path:
            return

        try:
            write_cargo_excel(self.cargo_registry, self.cargo_details, path)
        except Exception as exc:
            messagebox.showerror("내보내기 실패", f"엑셀로 내보내는 중 오류가 발생했습니다:\n{exc}")
            return

        messagebox.showinfo(
            "내보내기 완료",
            f"현재 화물 정보 {len(self.cargo_registry)}건을 저장했습니다:\n{path}",
        )

    def bulk_import_from_excel(self) -> None:
        """엑셀 파일을 통째로 읽어서 화물 등록부를 한 번에 채웁니다 (재배치 아님 - 단순 일괄 등록)."""
        path = filedialog.askopenfilename(
            title="화물 등록 엑셀 파일 선택",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if not path:
            return  # 사용자가 파일 선택 취소함

        try:
            new_registry, new_details, errors = bulk_import_cargo_from_excel(path, self.locations.keys())
        except Exception as exc:
            messagebox.showerror("가져오기 실패", f"엑셀 파일을 읽는 중 오류가 발생했습니다:\n{exc}")
            return

        if not new_registry:
            messagebox.showinfo("가져올 데이터 없음", "엑셀 파일에서 유효한 화물 데이터를 찾지 못했습니다.")
            return

        # 기존 등록부에 엑셀에서 읽은 내용을 덮어씀 (dict.update: 같은 키는 최신값으로 교체)
        self.cargo_registry.update(new_registry)
        self.cargo_details.update(new_details)
        save_cargo_registry(self.cargo_registry)
        save_cargo_details(self.cargo_details)
        for name, location in new_registry.items():
            _sync_cargo_to_db(name, location, new_details.get(name))
        self._refresh_cargo_list()

        summary = f"{len(new_registry)}건의 화물 위치를 등록했습니다."
        if errors:
            summary += "\n\n확인이 필요한 항목:\n" + "\n".join(errors[:15])  # 너무 길어지지 않게 15개까지만 표시
            if len(errors) > 15:
                summary += f"\n... 외 {len(errors) - 15}건"
        messagebox.showinfo("엑셀 일괄 등록 완료", summary)

    def _refresh_cargo_list(self) -> None:
        """Render the latest PostgreSQL inventory snapshot."""
        for widget in self.cargo_rows_container.winfo_children():
            widget.destroy()

        if self._inventory_records is None:
            message = (
                "DB 화물 정보를 불러오지 못했습니다."
                if self._inventory_error
                else "DB 화물 정보를 불러오는 중입니다…"
            )
            ctk.CTkLabel(self.cargo_rows_container, text=message,
                        font=self.font_body, text_color="#9e9e9e").pack(anchor="w", pady=10)
            return

        if not self._inventory_records:
            ctk.CTkLabel(self.cargo_rows_container, text="DB에 저장된 화물이 없습니다.",
                        font=self.font_body, text_color="#9e9e9e").pack(anchor="w", pady=10)
            return

        stale = bool(self._inventory_error)
        for cargo in self._inventory_records:
            row = ctk.CTkFrame(
                self.cargo_rows_container,
                fg_color="#2b2b2b",
                corner_radius=6,
                border_width=1,
                border_color="#c0392b" if stale else "#565b5e",
            )
            row.pack(fill="x", pady=(0, 8))
            row.grid_columnconfigure(0, weight=1)

            title = f"{cargo.name}  ·  ID {cargo.container_id or '-'}"
            location = f"위치 {cargo.location}  ·  {cargo.floor}층  ·  기반 ArUco {cargo.base_aruco_id or '-'}"
            details = f"종류 {cargo.cargo_type or '-'}"
            if cargo.note:
                details += f"  ·  비고 {cargo.note}"
            text_frame = ctk.CTkFrame(row, fg_color="transparent")
            text_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=8)
            ctk.CTkLabel(text_frame, text=title, font=("Malgun Gothic", 13, "bold"), text_color="#dce4ee", anchor="w").pack(fill="x", anchor="w")
            ctk.CTkLabel(text_frame, text=location, font=self.font_body, text_color="#92ccff", anchor="w").pack(fill="x", anchor="w", pady=(2, 0))
            ctk.CTkLabel(text_frame, text=details, font=self.font_body, text_color="#9e9e9e", anchor="w", wraplength=420, justify="left").pack(fill="x", anchor="w", pady=(2, 0))
            
            button_frame = ctk.CTkFrame(row, fg_color="transparent")
            button_frame.grid(row=0, column=1, padx=(0, 10), pady=8)
            ctk.CTkButton(
                button_frame,
                text="✏ 수정",
                width=58,
                fg_color="#343638",
                border_width=1,
                border_color="#565b5e",
                text_color="#f0ad4e",
                hover_color="#404040",
                command=lambda item=cargo: self.load_cargo_for_edit(item),
            ).pack(pady=(0, 4))
            action_frame = ctk.CTkFrame(row, fg_color="transparent")
            action_frame.grid(row=0, column=2, padx=(0, 10), pady=10)
            ctk.CTkButton(
                action_frame,
                text="🗣️ 배차",
                width=60,
                fg_color="#343638",
                border_width=1,
                border_color="#565b5e",
                text_color="#2e86c1",
                hover_color="#404040",
                command=lambda n=cargo.name: self.send_cargo_to_command_popup(n),
            ).pack(side="left", padx=(0, 4))
            ctk.CTkButton(
                action_frame,
                text="삭제",
                width=52,
                fg_color="#c0392b",
                hover_color="#922b21",
                text_color="white",
                state="disabled" if stale or self._inventory_delete_in_flight else "normal",
                command=lambda item=cargo: self._confirm_inventory_delete(item),
            ).pack(side="left")

    def load_cargo_for_edit(self, cargo) -> None:
        """Load one live DB row into the manual position editor."""
        container_id = str(cargo.container_id or '')
        if container_id not in MANUAL_CONTAINER_IDS:
            messagebox.showerror(
                '수정할 수 없는 ID',
                f'현재 시스템에서 지원하지 않는 컨테이너 ID입니다: {container_id or "-"}',
            )
            return
        location = str(cargo.location or '')
        if location not in MANUAL_INVENTORY_LOCATIONS:
            messagebox.showerror(
                '표준 위치 아님',
                f'현재 위치 {location or "-"}는 표준 운영 위치가 아닙니다.',
            )
            return
        self.container_id_var.set(container_id)
        self.location_option_var.set(location)
        self.floor_var.set(str(int(cargo.floor or 1)))

    def send_cargo_to_command_popup(self, name: str) -> None:
        """화물 목록에서 이 버튼을 누르면, 그 화물명이 미리 채워진 채로 명령 팝업이 열립니다.
        목적지만 이어서 입력하고 실행하면 됩니다 (예: "화물A를 " + "항구로 옮겨")."""
        from command_center import open_command_popup
        open_command_popup(self, on_cargo_updated=self.refresh_from_disk, initial_text=f"{name}를 ")

    # ------------------------------------------------------------------
    # 엑셀 기반 일괄 재배치
    # ------------------------------------------------------------------
    def import_relocation_excel(self) -> None:
        """엑셀(바뀐 위치들)을 불러와서 기존 등록부와 비교(diff)한 뒤,
        실제로 위치가 바뀐 화물들만 골라 자동으로 이동 경로를 계획하고 큐에 등록합니다."""
        path = filedialog.askopenfilename(
            title="재배치 엑셀 파일 선택",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if not path:
            return

        try:
            jobs, unchanged, new_items, new_registry, new_details, errors, final_next_index = plan_relocations_from_excel(
                path, self.cargo_registry, self.locations.keys(),
                vehicle_positions=self.vehicle_positions, waypoint_rules=self.rules,
            )
        except Exception as exc:
            messagebox.showerror("가져오기 실패", f"엑셀 파일을 읽는 중 오류가 발생했습니다:\n{exc}")
            return

        # 차량 순번(다음 배차 순서)은 이번 배치 계획에 맞춰 바로 반영해둡니다.
        # (차량의 실제 위치는 큐가 한 단계씩 진행되면서 advance_queue에서 갱신됩니다)
        self.next_vehicle_index = final_next_index

        # 세부정보(컨테이너/ArUco ID 등)는 항상 최신으로 갱신 - 위치 자체는 큐 실행 완료 시에만 갱신
        self.cargo_details.update(new_details)
        save_cargo_details(self.cargo_details)

        # 처음 등록되는 화물은 이동할 이전 위치가 없으므로 바로 등록
        for name in new_items:
            self.cargo_registry[name] = new_registry[name]
        if new_items:
            save_cargo_registry(self.cargo_registry)
            self._refresh_cargo_list()

        if not jobs:
            self.relocation_label.configure(
                text=f"이동이 필요한 화물이 없습니다. (변경없음 {len(unchanged)}건, 신규 등록 {len(new_items)}건)",
                text_color="gray70",
            )
            return

        lines = []
        for job in jobs:
            route_text = " → ".join(f"{s.location}[{s.action}]" for s in job["route"])
            container_note = f" (ArUco: {job['container_id']})" if job.get("container_id") else ""
            lines.append(f"- {job['item']}{container_note}: {route_text}")

        summary = (
            f"이동 필요 {len(jobs)}건 / 변경없음 {len(unchanged)}건 / 신규 등록 {len(new_items)}건\n"
            + "\n".join(lines)
        )
        self.relocation_label.configure(text=summary, text_color="#28C76F")

        self._log(f"[재배치 계획] 엑셀 불러옴 - 이동 필요 {len(jobs)}건, 변경없음 {len(unchanged)}건, 신규 {len(new_items)}건")
        for line in lines:
            self._log(line)
        if errors:
            self._log(f"[확인 필요] {len(errors)}건 - 상세는 엑셀 일괄 등록 시 표시되는 목록 참고")

        self._start_queue(jobs)

    # ------------------------------------------------------------------
    # 큐 기반 배차 실행 (엑셀 재배치 다건 처리용 - 자연어 명령은 command_center.py 참고)
    # ------------------------------------------------------------------
    def _start_queue(self, jobs: List[Dict]) -> None:
        """새로운 작업 목록으로 큐를 초기화합니다. (자연어 명령의 job 1개짜리든,
        엑셀 재배치의 job 여러 개짜리든 이 함수 하나로 통일해서 시작합니다)"""
        self.dispatch_queue = jobs
        self.active_job_index = 0   # 큐의 첫 번째 화물부터 시작
        self.active_step_index = 0  # 그 화물 경로의 첫 번째 단계부터 시작
        self.step_status_label.configure(text="")

        if jobs:
            first_label = jobs[0].get("label") or jobs[0].get("item")
            self.step_button.configure(text=f"배차 실행 (데모 시작 - {first_label})", state="normal")
        else:
            self.step_button.configure(state="disabled")

    def advance_queue(self) -> None:
        """"다음 단계로" 버튼을 누를 때마다 호출됩니다. 큐에 쌓인 화물들을 하나씩,
        그 화물의 경로도 한 단계씩 순서대로 진행시키는 상태 머신입니다.

        진행 흐름:
          1. 지금 큐에서 처리 중인 화물(job)과, 그 화물 경로에서 지금 단계(step)를 가져옴
          2. 로그를 남기고 화면 상태 텍스트 갱신
          3. 이번 화물의 마지막 단계까지 끝났으면:
             - 일반 화물 job이면 화물 위치를 최종 도착지로 갱신하고 저장
             - 차량 복귀 job(is_return)이면 화물 등록부는 건드리지 않고 차량 위치만 갱신
             - 전체 경로를 한 문장으로 요약해서 로그에 남김
             - 다음 화물(job)로 넘어감 (더 없으면 전체 완료 처리)
             - 아직 남았으면: 다음 단계로 버튼 텍스트만 갱신
        """
        if not self.dispatch_queue or self.active_job_index >= len(self.dispatch_queue):
            return  # 큐가 비어있거나 이미 다 끝난 상태면 아무것도 안 함

        job = self.dispatch_queue[self.active_job_index]
        route: List[RouteStep] = job["route"]
        label = job.get("label") or job.get("item") or "AGV"
        step = route[self.active_step_index]

        if self.active_step_index == 0:
            self._log(f"[{label}] 이동 시작 - {step.location} 에서 출발")
        else:
            prev = route[self.active_step_index - 1]
            self._log(f"[{label}] 이동 중: {prev.location} → {step.location} ({step.action})")
            self._log(f"[{label}] 도착: {step.location} - {step.action} 단계 진행")

        job_progress = f"[{self.active_job_index + 1}/{len(self.dispatch_queue)}]"  # 예: "[2/5]"
        self.step_status_label.configure(text=f"{job_progress} {label} - 현재 단계: {step.location} ({step.action})")

        self.active_step_index += 1  # 다음 클릭을 위해 미리 한 칸 전진

        if self.active_step_index >= len(route):
            item = job.get("item")
            vehicle_idx = job.get("vehicle_index")

            if job.get("is_return"):
                # 차량 복귀 job - 화물 등록부는 안 건드리고 차량 위치만 갱신
                if vehicle_idx is not None:
                    self.vehicle_positions[vehicle_idx] = route[-1].location
                self._log(f"[{label}] 복귀 완료 - 대기장소 도착")
            else:
                # 일반 화물 job - 화물 위치를 최종 목적지로 갱신
                self.cargo_registry[item] = route[-1].location
                save_cargo_registry(self.cargo_registry)
                self._refresh_cargo_list()
                if vehicle_idx is not None:
                    self.vehicle_positions[vehicle_idx] = route[-1].location
                self._log(f"[{label}] 완료 - 위치가 '{route[-1].location}'로 갱신됨")

            self._log(describe_route_sentence(label, route))
            save_vehicle_state(self.vehicle_positions, self.next_vehicle_index)

            self.active_job_index += 1   # 큐의 다음 화물로 이동
            self.active_step_index = 0   # 새 화물이니 단계는 처음부터

            if self.active_job_index >= len(self.dispatch_queue):
                # 큐에 더 이상 화물이 없으면 전체 작업 종료
                self._log("[전체 완료] 모든 배차 작업이 끝났습니다.")
                self.step_button.configure(text="배차 완료", state="disabled")
            else:
                # 다음 화물이 남아있으면 버튼 텍스트만 바꿔서 계속 누를 수 있게 함
                next_job = self.dispatch_queue[self.active_job_index]
                next_label = next_job.get("label") or next_job.get("item") or "AGV"
                self.step_button.configure(text=f"다음 화물로 ({next_label})")
        else:
            # 같은 화물의 다음 단계가 남아있는 경우
            next_step = route[self.active_step_index]
            self.step_button.configure(text=f"다음 단계로 ({next_step.location})")

    # ------------------------------------------------------------------
    def _log(self, text: str) -> None:
        """진행 로그 창에 시각 태그를 붙여 한 줄 추가합니다."""
        self.log_box.configure(state="normal")  # 잠깐 편집 가능 상태로 풀었다가
        timestamp = time.strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {text}\n")
        self.log_box.see("end")  # 항상 최신 로그가 보이도록 맨 아래로 스크롤
        self.log_box.configure(state="disabled")  # 다시 읽기 전용으로 잠금


# -----------------------------------------------------------------------------
# 실차(ROS2 Nav2) 연동 참고 코드
# -----------------------------------------------------------------------------
#
# 실제 AGV와 연결할 때는 위 GUI의 "다음 단계로" 버튼 클릭을 아래와 같은 자동화된
# 흐름으로 바꾸면 됩니다. rclpy/nav2_msgs가 있는 실제 워크스페이스에서만 동작하며,
# 이 환경에는 ROS2가 없어 실행 테스트는 못 했습니다 - 실제 워크스페이스에서 한 번
# 돌려보시고 에러 있으면 바로 고쳐드리겠습니다.
#
# from location_to_nav_goal import send_nav_goal
#
# class CargoDispatcherROS2:
#     def __init__(self, lookup_path: str):
#         self.lookup_path = lookup_path
#
#     def dispatch(self, item_name: str, destination: str, cargo_registry: dict):
#         route = build_route(item_name, destination, cargo_registry)
#         for step in route:
#             send_nav_goal(step.location, self.lookup_path)   # 이동 완료까지 블로킹
#             if step.action == "적재":
#                 self.wait_for_load_confirmation()             # 적재 확인 대기
#             elif step.action == "하역":
#                 self.wait_for_unload_confirmation()            # 하역 확인 대기
#         cargo_registry[item_name] = route[-1].location
#
#     def wait_for_load_confirmation(self):
#         # TODO: 중앙 카메라가 "차량 위 화물 적재됨"을 발행하는 토픽 구독 후
#         #       메시지가 올 때까지 대기하도록 구현 (예: rclpy.spin_until_future_complete)
#         raise NotImplementedError
#
#     def wait_for_unload_confirmation(self):
#         # TODO: 하역 완료 확인 토픽 구독 대기
#         raise NotImplementedError


# -----------------------------------------------------------------------------
# 단독 실행
# -----------------------------------------------------------------------------

if __name__ == "__main__":
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.title("화물 위치 추적 및 엑셀 재배치")
    root.geometry("1100x700")

    app = CargoDispatchTool(root)
    app.pack(fill="both", expand=True, padx=15, pady=15)

    root.mainloop()
